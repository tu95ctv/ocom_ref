from datetime import datetime
import os
from typing import Type
import pandas as pd

path = r'C:\d4\sonix\a_duc_bidcrawler\bidcrawler\webbc\assets\cache'
list_of_files = {}
for (dirpath, dirnames, filenames) in os.walk(path):
    for filename in filenames:
        if 'git' not in filename:
	        list_of_files[filename] = os.sep.join([dirpath, filename])

# print (list_of_files)
## code gốc
# from core.utils import file_helpers
# for path in list_of_files.values():
#     print (path)
#     ret, content = file_helpers.read_msexcel(path)
#     print (bool(content))


#just open file
import pandas as pd
import xlrd
# for path in list_of_files.values():
#     print (path)
#     try:
#         wb = xlrd.open_workbook( path )
#         # df = pd.read_excel(path, sheet_name=None, header=None)
#         # df = _process(path)
#     except xlrd.biffh.XLRDError as e:# lỗi mật khẩu
#         print ('lỗi mật khẩu: %s'%e)
#         pass

path = 'C:\d4\sonix\a_duc_bidcrawler\bidcrawler\webbc\assets\cache\93268e6262a0480baa458304e6d5099a.xlsx'
path = r'C:\d4\sonix\a_duc_bidcrawler\bidcrawler\webbc\assets\cache\0cfd0c60030d4a1f9855e8510961dec8.xlsx'
path = r'C:\d4\sonix\a_duc_bidcrawler\bidcrawler\webbc\assets\cache\0f3e6e4976504272aad431e307cf0870.xlsx'
path = r'C:\d4\sonix\a_duc_bidcrawler\bidcrawler\webbc\assets\cache\abc.xlsx'
path = r'C:\d4\sonix\a_duc_bidcrawler\bidcrawler\webbc\assets\cache\29f81ba15fc047939b704b488b3598f0.xlsx'
path = r'C:\Users\nguye\Downloads\3-r2sinsyou-sityousonbetu-syougaibetu.xlsx'#charset#pandas thì báo lỗi, còn openpyxl thì không
# path = r'C:\Users\nguye\Downloads\zogenhoka.xlsx'#charset
# path = r'C:\Users\nguye\Downloads\R3minibasketball_moushikomi.xlsx'#max value is 14, bị lỗi ngay khi wb_obj = openpyxl.load_workbook(path)
# path = r'C:\Users\nguye\Downloads\131206.xlsx'#max value is 14
# # path = r'C:\Users\nguye\Downloads\h26.xlsx'#Element
# path = r'C:\Users\nguye\Downloads\R01toukei_nendo.xlsx'# out of range, loi raise o openpyxl.load_workbook
# # path = r'C:\Users\nguye\Downloads\r03_youshiki01.xlsx'
# path = r'C:\Users\nguye\Downloads\r03_youshiki02.xlsx'
# path = r'C:\Users\nguye\Downloads\R01sankou_nendo.xlsx'
# path = r'C:\Users\nguye\Downloads\2904-4.xlsx'# __init__, nhu loi index khong fix duodcj

    
#đoạn code của anh Đức đọc file
def _process(filename):
    import xlwings as xw
    wb = xw.Book(filename)
    sheet = wb.sheets[0]
    df = sheet.used_range.options(pd.DataFrame,  header=False).value
    wb.close()
    return df

def read_one_file_xlrd(path):
    workbook = xlrd.open_workbook( path )
    print ('*workbook.sheets()*', workbook.sheets())
    output = ''
    for sheet_count, sheet in enumerate(workbook.sheets()):
        output += f'Sheet name {sheet.name}' + "\n"
        # try:
        sheets= []
        for row in range(sheet.nrows):
            rows = []
            for col in range(sheet.ncols):
                val = sheet.cell_value(row,col)
                rows.append(str(val))
            sheets.append(rows)
        strsheets =  '\n'.join([' '.join(cols) for cols in sheets])
        output +=strsheets + ('\n' if sheet_count != len(workbook.sheets()) else '')
    print ('**output**', output)
    return output
        # except Exception as err:
        #     print (err)

    
def read_one_file_pd(path):
    from datetime import datetime
    content = ''
    df = pd.read_excel(path, sheet_name=None, header=None)
    # df = _process(path)
    for key, value in df.items():
        data_list = value.fillna('').values.tolist()
        for row in data_list:
            for col in row:
                if isinstance(col, datetime):
                    try:
                        row[row.index(col)] = col.strftime("%Y/%m/%d")
                    except:
                        row[row.index(col)] = str(col)
                else:
                    row[row.index(col)] = str(col)
            content += ' '.join(row) + '\n'
    return content



def read_one_file_process(path):
    from datetime import datetime
    content = ''
    df = _process(path)
    for key, value in df.items():
        print ('key', key)
        data_list= [str(i) for i in value]
        content += ' '.join(data_list) + '\n'
    return content


import openpyxl

def convert(v):
    if isinstance(v,datetime):
        v = v.strftime('%d-%m-%Y')
        return v
    elif v == None:
        return ''
    else:
        return str(v)


def read_one_file_openpyxl(path):
    from datetime import datetime
    content = ''
    content
    # df = pd.read_excel(path, sheet_name=None, header=None)
    wb_obj = openpyxl.load_workbook(path)
    print ('wb_obj.worksheets',wb_obj.worksheets)
    # sheet_obj = wb_obj.active
    # max_col = sheet_obj.max_column
    # print ('**max_col', type(sheet_obj))


# đọc 1 file
# content = read_one_file(path)
# content = read_one_file_process(path)
content = read_one_file_pd(path)
# content = read_one_file_xlrd(path)
# content  = read_one_file_openpyxl(path)
# 
# print ('**content**', content)
# # đọc 1 file
#
# # df = pd.read_excel(path, sheet_name=None, header=None)
# df = _process(path)
# for key, value in df.items():
#     print ('key',key)
#     print ('value',value, 'type value', type(value))
#     data_list = value.fillna('').values.tolist()
#     print ('**data_list', data_list)
    # for row in data_list:
    #     for col in row:
            # print ('row',row, type(row))
            # row[row.index(col)] = str(col)
        # content += ' '.join(row) + '\n'







